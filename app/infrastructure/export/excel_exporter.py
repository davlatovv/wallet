import calendar
import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.domain.entities.transaction import TransactionType
from app.domain.repositories.abstract_transaction import AbstractTransactionRepository

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2E86AB")
_ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
_SUMMARY_FONT = Font(bold=True)
_SUMMARY_FILL = PatternFill("solid", fgColor="D9EAD3")

_TX_TYPE_LABELS = {
    TransactionType.INCOME: "Доход",
    TransactionType.EXPENSE: "Расход",
    TransactionType.SAVINGS: "Копилка",
}


class ExcelExporter:
    def __init__(self, tx_repo: AbstractTransactionRepository) -> None:
        self._tx_repo = tx_repo

    async def export_month(self, user_id: int) -> bytes:
        now = datetime.now(timezone.utc)
        return await self.export_by_month(user_id, now.year, now.month)

    async def export_by_month(self, user_id: int, year: int, month: int) -> bytes:
        from_dt = datetime(year, month, 1, tzinfo=timezone.utc)
        last_day = calendar.monthrange(year, month)[1]
        to_dt = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)
        transactions = await self._tx_repo.list_by_period(user_id, from_dt, to_dt)

        wb = Workbook()
        ws = wb.active
        ws.title = "Транзакции"

        headers = ["Дата", "Тип", "Сумма (UZS)", "Категория", "Заметка"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        total_income = total_expense = total_savings = 0.0
        for row_idx, tx in enumerate(transactions, 2):
            amount = float(tx.amount)
            fill = _ALT_FILL if row_idx % 2 == 0 else None

            date_cell = ws.cell(row=row_idx, column=1, value=tx.created_at.strftime("%d.%m.%Y %H:%M"))
            type_cell = ws.cell(row=row_idx, column=2, value=_TX_TYPE_LABELS.get(tx.transaction_type, ""))
            amount_cell = ws.cell(row=row_idx, column=3, value=amount)
            amount_cell.number_format = '#,##0.00'
            cat_cell = ws.cell(row=row_idx, column=4, value=tx.category_name or "")
            note_cell = ws.cell(row=row_idx, column=5, value=tx.note or "")

            if fill:
                for cell in (date_cell, type_cell, amount_cell, cat_cell, note_cell):
                    cell.fill = fill

            if tx.transaction_type == TransactionType.INCOME:
                total_income += amount
            elif tx.transaction_type == TransactionType.EXPENSE:
                total_expense += amount
            elif tx.transaction_type == TransactionType.SAVINGS:
                total_savings += amount

        # Summary block
        summary_row = len(transactions) + 3
        summary_items = [
            ("Итого доходов:", total_income),
            ("Итого расходов:", total_expense),
        ]
        if total_savings > 0:
            summary_items.append(("В копилку:", total_savings))
        summary_items.append(("Баланс:", total_income - total_expense - total_savings))

        for i, (label, value) in enumerate(summary_items):
            label_cell = ws.cell(row=summary_row + i, column=4, value=label)
            value_cell = ws.cell(row=summary_row + i, column=5, value=value)
            value_cell.number_format = '#,##0.00'
            label_cell.font = _SUMMARY_FONT
            value_cell.font = _SUMMARY_FONT
            label_cell.fill = _SUMMARY_FILL
            value_cell.fill = _SUMMARY_FILL

        col_widths = [18, 12, 16, 20, 30]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
